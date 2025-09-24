"""Generator rapoarte ANRE in format Excel."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


HEADER_MAP = {
    "energie_activa_produsa": "Energie activa produsa (MWh)",
    "energie_activa_livrata": "Energie activa livrata (MWh)",
    "energie_consum_propriu": "Consum propriu (MWh)",
    "energie_reactiva": "Energie reactiva (Mvarh)",
    "disponibilitate_centrala": "Disponibilitate (%)",
    "ore_functionare": "Ore functionare",
}


def generate_monthly_report(data: Dict[str, float], incidents: pd.DataFrame, output: Path, month: datetime) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport" if month.month else "Raport"
    title = f"Raport ANRE {month.strftime('%B %Y')}"
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal="center")

    row = 3
    for key, label in HEADER_MAP.items():
        ws[f"A{row}"] = label
        ws[f"B{row}"] = data.get(key, 0.0)
        row += 1

    ws[f"A{row+1}"] = "Incidente raportate (conform Ordin ANRE 177/2015)"
    incident_row = row + 2
    ws.append(["Data", "Durata (h)", "Descriere", "Impact"],)
    header_font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=incident_row, column=col).font = header_font
    for _, incident in incidents.iterrows():
        incident_row += 1
        ws.append(
            [
                incident.get("data"),
                incident.get("durata_ore"),
                incident.get("descriere"),
                incident.get("impact"),
            ]
        )
    wb.save(output)
